"""generate_test_suite: spec -> .xlsx suite, and the generated sheet round-trips the parser.

Offline: pure file I/O, no network. The key contract is that ``read_test_suite`` parses the
generated sheet with zero parse errors, so it is immediately runnable by ``run_and_record``.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from api_log_test_mcp.tools.suite import read_test_suite
from api_log_test_mcp.tools.suite_generator import generate_test_suite

SPEC = Path(__file__).parent.parent / "resources" / "products-eapi1.yaml"


@pytest.fixture
def generated(tmp_path: Path) -> tuple[dict, str]:
    out = tmp_path / "products_suite.xlsx"
    summary = generate_test_suite(str(SPEC), str(out))
    return summary, str(out)


def test_writes_sheet_with_spec_basepath(generated):
    summary, out = generated
    assert Path(out).exists()
    assert summary["base_path"] == "https://api.example.com/products/v1"
    assert summary["case_count"] >= 30  # comprehensive coverage


def test_worksheet_is_named_tests(generated):
    from openpyxl import load_workbook

    _, out = generated
    wb = load_workbook(out)
    assert wb.sheetnames == ["tests"]


def test_round_trips_through_parser_without_errors(generated):
    summary, out = generated
    suite = read_test_suite(out)
    assert suite.base_path == summary["base_path"]
    assert suite.parse_errors == []
    assert len(suite.cases) == summary["case_count"]


def test_covers_every_validation_category(generated):
    _, out = generated
    suite = read_test_suite(out)
    statuses = {c.expected_status for c in suite.cases}
    # positive + body(422) + query/path/bad-request(400) + media-type(415) + 404 + auth(401)
    assert {200, 201, 400, 401, 404, 415, 422} <= statuses

    descriptions = " || ".join(c.description or "" for c in suite.cases)
    assert "missing required" in descriptions  # required-field negative
    assert "violates pattern" in descriptions  # pattern negative
    assert "not in allowed enum" in descriptions  # enum negative
    assert "exceeds maxItems" in descriptions  # array negative


def test_malformed_body_case_kept_as_raw_string(generated):
    _, out = generated
    suite = read_test_suite(out)
    malformed = [c for c in suite.cases if c.expected_status == 400 and c.method == "POST"]
    assert malformed, "expected a malformed-JSON POST case"
    # The deliberately-broken body survives as a raw scalar string (not valid JSON).
    assert isinstance(malformed[0].body, str)
